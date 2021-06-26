import openpyxl
from openpyxl import load_workbook

workbook = openpyxl.load_workbook(filename='datos.xlsx')
worksheet = workbook['Hoja1']
nrows = len(worksheet['A'])
worksheet.cell(row=nrows+1, column=1, value = 'Maria')
worksheet.cell(row=nrows+1, column=2, value = '21')
worksheet.cell(row=nrows+1, column=3, value = 'No')
workbook.save('datos.xlsx')

print(nrows)
